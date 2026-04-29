"""
Loose Number Extractor
Streamlit app — deployable to Streamlit Community Cloud
"""

import io
import re
import difflib
import datetime as dt
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st

# ─── Page config ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Loose Number Extractor",
    page_icon="📄",
    layout="centered",
)

# ─── PDF helpers ──────────────────────────────────────────────────────────────

def extract_pdf_pages(pdf_file):
    pages = []
    with pdfplumber.open(pdf_file) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            tables = page.extract_tables() or []
            pages.append((i, text, tables))
    return pages


def split_sentences(text):
    parts = re.split(r'(?<=[.!?])\s+(?=[A-Z\$\d])|(?:\n\s*\n)', text)
    return [p.strip() for p in parts if p.strip()]


_STOPWORDS = {
    "the", "and", "for", "that", "with", "this", "from", "are", "was",
    "were", "have", "has", "been", "each", "only", "also", "such",
    "loan", "loans", "mortgage", "mortgages", "date", "dates",
}


def _is_page_ref(text, match_start):
    prefix = text[max(0, match_start - 30): match_start]
    if re.search(r'\.{3,}\s*$', prefix):
        return True
    if re.search(r'\bpage\s+$', prefix, re.I):
        return True
    if re.search(r'\bp\.\s*$', prefix, re.I):
        return True
    if re.search(r'\b(?:section|exhibit|appendix|schedule|annex|figure)\s+$', prefix, re.I):
        return True
    return False


def extract_numbers(text):
    results = []

    for m in re.finditer(r'\$[\d,]+(?:\.\d+)?', text):
        try:
            results.append(("dollar", float(m.group().replace("$", "").replace(",", "")), m.group()))
        except ValueError:
            pass

    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*%', text):
        try:
            results.append(("pct", float(m.group(1)), f"{m.group(1)}%"))
        except ValueError:
            pass

    for m in re.finditer(r'(?<![a-zA-Z\d])\((\d{1,7})\)', text):
        try:
            val = int(m.group(1))
            if val <= 20:
                continue
            results.append(("count", val, m.group()))
        except ValueError:
            pass

    for m in re.finditer(r'(?<!\$)\b(\d{1,3}(?:,\d{3})+)\b', text):
        try:
            results.append(("number", int(m.group().replace(",", "")), m.group()))
        except ValueError:
            pass

    for m in re.finditer(r'\b(\d{1,4})\s+(?:months?|basis points?|bps?)\b', text, re.I):
        try:
            results.append(("count", int(m.group(1)), m.group(1)))
        except ValueError:
            pass

    for m in re.finditer(r'(?<!\d)(\d{3,6})(?!\d)(?!%)', text):
        try:
            val = int(m.group(1))
            if 2000 <= val <= 2099:
                continue
            if _is_page_ref(text, m.start()):
                continue
            results.append(("integer", val, m.group(1)))
        except ValueError:
            pass

    return results


def _pct_to_decimal(val):
    return val / 100


def primary_number(text):
    if re.search(r'\[\s*\]', text):
        return "[ ]"
    nums = extract_numbers(text)
    if not nums:
        return None
    for ntype in ("dollar", "pct", "count", "number", "integer"):
        found = [n for n in nums if n[0] == ntype]
        if found:
            return _pct_to_decimal(found[0][1]) if ntype == "pct" else found[0][1]
    return None


def _raw_precision(raw_str):
    """Count decimal places in a raw number string."""
    s = re.sub(r'[$,]', '', raw_str.strip().rstrip('%'))
    if '.' in s:
        return len(s.split('.')[1])
    return 0


def extract_ordered_numbers(text):
    """Return list of (value, precision) tuples from text in order of appearance.
    Precision = decimal places of the value as stored in col F.
    For percentages stored as decimals, precision = pct_decimal_places + 2.
    """
    candidates = []

    for m in re.finditer(r'\$[\d,]+(?:\.\d+)?', text):
        try:
            raw = m.group()
            val = float(raw.replace('$', '').replace(',', ''))
            prec = _raw_precision(raw)
            candidates.append((m.start(), m.end(), val, prec))
        except ValueError:
            pass

    for m in re.finditer(r'(\d+(?:\.\d+)?)\s*%', text):
        try:
            raw_pct = m.group(1)
            val = float(raw_pct) / 100
            pct_prec = len(raw_pct.split('.')[1]) if '.' in raw_pct else 0
            prec = pct_prec + 2
            candidates.append((m.start(), m.end(), val, prec))
        except ValueError:
            pass

    for m in re.finditer(r'(?<!\$)(\d{1,3}(?:,\d{3})+)', text):
        try:
            val = int(m.group().replace(',', ''))
            candidates.append((m.start(), m.end(), val, 0))
        except ValueError:
            pass

    for m in re.finditer(r'(\d{1,4})\s+(?:months?|basis\s*points?|bps?)', text, re.I):
        try:
            val = int(m.group(1))
            candidates.append((m.start(), m.end(), val, 0))
        except ValueError:
            pass

    for m in re.finditer(r'(?<!\d)(\d{3,6})(?!\d)(?!%)', text):
        try:
            v = int(m.group(1))
            if 2000 <= v <= 2099:
                continue
            candidates.append((m.start(), m.end(), v, 0))
        except ValueError:
            pass

    candidates.sort(key=lambda x: x[0])
    result = []
    used_end = -1
    for start, end, val, prec in candidates:
        if start >= used_end:
            result.append((val, prec))
            used_end = end
    return result


def contextual_number(template_lang, pdf_text):
    tl = template_lang.lower().strip()

    day_match = re.search(r'(\d+)\s+(?:to\s+\d+|or\s+more)', tl)
    if day_match:
        pattern = re.escape(day_match.group(0))
        m = re.search(pattern, pdf_text, re.I)
        if m:
            nearby = pdf_text[m.start(): m.start() + 200]
            nums = extract_numbers(nearby)
            for ntype in ("pct", "dollar", "count", "number", "integer"):
                found = [n for n in nums if n[0] == ntype]
                if found:
                    v = found[0][1]
                    return _pct_to_decimal(v) if ntype == "pct" else v

    nums = extract_numbers(pdf_text)

    if tl == "low":
        for ntype in ("pct", "dollar", "number", "integer"):
            found = [n for n in nums if n[0] == ntype]
            if found:
                v = min(found, key=lambda n: n[1])[1]
                return _pct_to_decimal(v) if ntype == "pct" else v

    if tl == "high":
        for ntype in ("pct", "dollar", "number", "integer"):
            found = [n for n in nums if n[0] == ntype]
            if found:
                v = max(found, key=lambda n: n[1])[1]
                return _pct_to_decimal(v) if ntype == "pct" else v

    if re.search(r'\b(?:weighted\s+average|average|avg|w\.?a\.?)\b', tl):
        for ntype in ("pct", "dollar", "number", "integer"):
            found = [n for n in nums if n[0] == ntype]
            if found:
                v = found[-1][1]
                return _pct_to_decimal(v) if ntype == "pct" else v

    return primary_number(pdf_text)


_TOC_RE = re.compile(
    r'(?:^\d{1,3}\s+\S.*\.{3}'
    r'|\.{4,}\s*\d{1,3}\s*$)',
    re.MULTILINE,
)


def gather_number_sentences(pdf_pages):
    seen = set()
    rows = []

    def add(page_num, chunk, label_only=False):
        chunk = chunk.strip()
        if len(chunk) < 4:
            return
        if _TOC_RE.match(chunk):
            return
        has_placeholder = bool(re.search(r'\[\s*\]', chunk))
        nums = extract_numbers(chunk)
        if not nums and not has_placeholder:
            return
        key = chunk[:100].lower()
        if key in seen:
            return
        seen.add(key)
        rows.append({"page": page_num, "language": chunk, "number": primary_number(chunk)})

    for page_num, text, tables in pdf_pages:
        for sent in split_sentences(text):
            add(page_num, sent)
        for line in text.splitlines():
            add(page_num, line.strip())
        for table in tables:
            for row in table:
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if len(cells) >= 2:
                    joined = "  ".join(cells)
                    if len(joined) <= 300:
                        add(page_num, joined)
                for cell in cells:
                    add(page_num, cell)

    return rows


# ─── Matching helpers ──────────────────────────────────────────────────────────

def keyword_overlap(a, b):
    wa = {w.lower() for w in re.findall(r'\b[a-zA-Z]{4,}\b', a) if w.lower() not in _STOPWORDS}
    wb = {w.lower() for w in re.findall(r'\b[a-zA-Z]{4,}\b', b) if w.lower() not in _STOPWORDS}
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def best_pdf_match(template_lang, pdf_rows, threshold=0.18):
    best_score, best = 0.0, None
    tl = template_lang.lower()
    for item in pdf_rows:
        pl = item["language"].lower()
        overlap = keyword_overlap(template_lang, item["language"])
        seq = difflib.SequenceMatcher(None, tl[:200], pl[:200]).ratio()
        starts_match = 0.0
        if len(template_lang) < 60:
            first_words = " ".join(tl.split()[:4])
            if pl.startswith(first_words):
                starts_match = 0.4
        score = 0.55 * overlap + 0.25 * seq + 0.20 * starts_match
        if score > best_score:
            best_score, best = score, item
    return (best, best_score) if best_score >= threshold else (None, best_score)


# ─── Excel helpers ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_WRAP        = Alignment(wrap_text=True, vertical="top")


def write_fresh_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loose Numbers"
    for col, label in [(5, "TS Language"), (6, "TS Boxed Numbers")]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, row in enumerate(rows, start=2):
        lang_cell = ws.cell(row=i, column=5, value=row.get("language", ""))
        lang_cell.alignment = _WRAP
        num = row.get("number")
        if num is not None:
            ws.cell(row=i, column=6, value=num)
    ws.column_dimensions["E"].width = 90
    ws.column_dimensions["F"].width = 22
    return wb


def write_updated_excel(template_wb, updated_rows):
    ws = template_wb.active
    for entry in updated_rows:
        r = entry["row_idx"]
        if entry.get("language") is not None:
            cell = ws.cell(row=r, column=5, value=entry["language"])
            cell.alignment = _WRAP
        new_val = entry.get("number")
        if new_val is not None:
            orig_val = ws.cell(row=r, column=6).value
            if isinstance(orig_val, dt.datetime) and not isinstance(new_val, dt.datetime):
                pass
            else:
                ws.cell(row=r, column=6, value=new_val)
        prec = entry.get("precision")
        if prec is not None:
            ws.cell(row=r, column=7, value=f"=ROUND(C{r},H{r})-F{r}")
            ws.cell(row=r, column=8, value=prec)
        else:
            ws.cell(row=r, column=7, value="Handtie")
    ws.column_dimensions["E"].width = 90
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 12
    return template_wb


def load_template_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append({
            "row_idx":  r,
            "language": ws.cell(row=r, column=5).value,
            "number":   ws.cell(row=r, column=6).value,
        })
    return wb, rows


# ─── Streamlit UI ──────────────────────────────────────────────────────────────

st.markdown("""
<style>
    .block-container { max-width: 720px; }
    .stDownloadButton > button {
        background-color: #1F4E79;
        color: white;
        width: 100%;
        font-size: 1rem;
        font-weight: 700;
        padding: 0.6rem;
        border-radius: 8px;
    }
    .stDownloadButton > button:hover { background-color: #163d5e; }
</style>
""", unsafe_allow_html=True)

st.title("📄 Loose Number Extractor")
st.caption("Extract term sheet & PPM numbers & language into Excel")

st.info(
    "**Two modes:**\n"
    "- **PDF only** — extracts every sentence with a number into a fresh Excel (columns E & F).\n"
    "- **PDF + previous Excel** — matches each row's existing language to the new PDF and "
    "updates columns E, F, G (Etie formula), and H (Precision) in place.",
    icon="ℹ️",
)

st.divider()

pdf_file   = st.file_uploader("New Term Sheet / PPM PDF  *(required)*", type=["pdf"])
excel_file = st.file_uploader("Previous Excel File  *(optional — enables update mode)*", type=["xlsx", "xls"])

st.divider()

if st.button("⚡ Extract & Download Excel", type="primary", use_container_width=True):
    if not pdf_file:
        st.error("Please upload a PDF file first.")
    else:
        with st.spinner("Extracting numbers from PDF…"):
            try:
                pdf_pages = extract_pdf_pages(pdf_file)
                pdf_rows  = gather_number_sentences(pdf_pages)

                if excel_file:
                    template_wb, template_rows = load_template_excel(excel_file)
                    updated = []
                    _CONTEXT_LABELS = {"low", "high", "avg", "average", "wa", "w.a."}
                    last_match_chunk = None
                    chunk_cursors = {}

                    for row in template_rows:
                        lang = row["language"]
                        if not lang or not isinstance(lang, str) or not lang.strip():
                            updated.append({"row_idx": row["row_idx"], "language": None, "number": None, "precision": None})
                            continue

                        lang_stripped = lang.strip()
                        tl = lang_stripped.lower()

                        if tl in _CONTEXT_LABELS and last_match_chunk:
                            num = contextual_number(lang_stripped, last_match_chunk)
                            updated.append({"row_idx": row["row_idx"], "language": None, "number": num, "precision": None})
                            continue

                        is_short = len(lang_stripped) <= 30
                        threshold = 0.35 if is_short else 0.18

                        match, score = best_pdf_match(lang_stripped, pdf_rows, threshold=threshold)
                        if match:
                            last_match_chunk = match["language"]
                            out_lang = lang_stripped if is_short else match["language"]

                            if re.search(r'\[\s*\]', match["language"]):
                                num = "[ ]"
                                prec = None
                            else:
                                ordered = extract_ordered_numbers(match["language"])
                                ck = match["language"][:200]
                                idx = chunk_cursors.get(ck, 0)
                                if ordered:
                                    num, prec = ordered[idx] if idx < len(ordered) else ordered[-1]
                                    chunk_cursors[ck] = idx + 1
                                else:
                                    num, prec = None, None

                            updated.append({
                                "row_idx":   row["row_idx"],
                                "language":  out_lang,
                                "number":    num,
                                "precision": prec,
                            })
                        else:
                            updated.append({
                                "row_idx":   row["row_idx"],
                                "language":  lang,
                                "number":    row["number"],
                                "precision": None,
                            })

                    out_wb   = write_updated_excel(template_wb, updated)
                    out_name = f"{Path(pdf_file.name).stem}_Updated.xlsx"
                else:
                    out_wb   = write_fresh_excel(pdf_rows)
                    out_name = f"{Path(pdf_file.name).stem}_Loose_Numbers.xlsx"

                buf = io.BytesIO()
                out_wb.save(buf)
                buf.seek(0)

                st.success("Done! Click below to download.")
                st.download_button(
                    label=f"⬇️  Download {out_name}",
                    data=buf,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as exc:
                st.error(f"Error: {exc}")

st.divider()
st.markdown(
    "<div style='text-align:center;color:#aaa;font-size:.8rem;'>"
    "Numbers extracted exactly as they appear in the PDF &nbsp;·&nbsp; "
    "Columns A–D left blank &nbsp;·&nbsp; Language → col E &nbsp;·&nbsp; Number → col F"
    "&nbsp;·&nbsp; Etie formula → col G &nbsp;·&nbsp; Precision → col H"
    "</div>",
    unsafe_allow_html=True,
)
